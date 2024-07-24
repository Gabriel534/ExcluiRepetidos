import openpyxl
import os
from pprint import pprint

PASTA = "Arquivos"

# Carregar o arquivo XLSX
workbook = openpyxl.load_workbook('CONTROLE.xlsx')

# Verificar as planilhas disponíveis
# print(workbook.sheetnames)

# Selecionar uma planilha específica (por exemplo, a primeira planilha)
ControleNotebooks = workbook["Controle notebooks"]
Componentes = workbook["Componentes"]
# Iterar pelas linhas e imprimir os valores de uma coluna
# print('Valores da coluna A:')
# for row in ControleNotebooks.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1):
#     for cell in row:
#         print(cell.value)
cont = 1
info = "AS"
info2 = "AS"
# modelos:list[str] = []
componentesLista:list[list[str], list[str]] = [[], []]

# while True:
#     info = ControleNotebooks[f"A{cont}"].value
#     if info == None: break
#     modelos.append(info)
#     cont += 1

cont = 1
while True:
    info = Componentes[f"A{cont}"].value
    componentesLista[0].append(info)
    info2 = Componentes[f"B{cont}"].value
    componentesLista[1].append(info2)
    
    if info == None: break
    if info2 == None: break
    cont += 1

if not os.path.exists(PASTA): os.mkdir(PASTA)

cont = 0

for modelo, componente in componentesLista[0], componentesLista[1]:
        if not os.path.exists(f"{PASTA}\\{modelo}"):
            print(f"Pasta {PASTA}\\{modelo} inexistente")
            continue
        print("Vish")

workbook.close()
